import openpyxl as excel
from openpyxl import Workbook
import numpy as np
from spire.xls import Workbook
from spire.xls import FileFormat

print("1つ目の比較したいExcelファイル名を入力してください：")
forst_file = input()

Fbook = excel.load_workbook(forst_file)
Fsheet = Fbook.active

flist = np.array([cell.value for cell in Fsheet['D'][1:]])
flist_beside = flist.reshape(len(x), 1)

fColumn_tmp = np.hstack([flist_beside])
fColumn = fColumn_tmp.tolist()
# print(fColumn)

print("2つ目の比較したいExcelファイル名を入力してください：")
second_file = input()

Sbook = excel.load_workbook(second_file)
Ssheet = Sbook.active

slist1 = np.array([cell.value for cell in Ssheet['A'][1:]])
slist2 = np.array([cell.value for cell in Ssheet['B'][1:]])

slist1_beside = x.reshape(len(x), 1)
slist2_beside = y.reshape(len(y), 1)

sColumn_tmp = np.hstack([x_beside, y_beside])
sColumn = sColumn_tmp.tolist()
# print(sColumn)

tmp_list = []
for i in range(len(fColumn)):
    for j in range(len(sColumn)):
        if fColumn[i][0] == sColumn[j][1]:
            tmp_list.append([sColumn[j][0]])
            break

tmp_list = np.array(tmp_list)
tmp_list_beside = tmp_list.reshape(len(tmp_list), 1)
tmp_list_beside = tmp_list_beside.tolist()
# print(tmp_list_beside)

workbook = Workbook()

workbook.Worksheets.Clear()
worksheet = workbook.Worksheets.Add("抽出された列")

i = 1
for array in tmp_list_beside:
    worksheet.InsertArray(array, i, 3, False)
    i += 1

workbook.SaveToFile("output/配列をシートに書き込む.xlsx", FileFormat.Version2016)
workbook.Dispose() 
